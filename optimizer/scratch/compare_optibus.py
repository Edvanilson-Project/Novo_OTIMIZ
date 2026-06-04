"""Real OTIMIZ-vs-Optibus comparison on two production instances.

Loads Optibus '_full_schedule' xlsx exports, extracts the underlying timetable
(service_trip rows, segments merged into real trips), runs ALL registered VSP
algorithms + the full hybrid pipeline (for crew duties), and compares
vehicles/duties/cost against Optibus's own solution embedded in the same file.

Calibration is derived from Optibus's own schedule so the comparison is fair:
  - min_layover_minutes: Optibus chains back-to-back (median gap 0)
  - max_block_span_minutes=1440: Optibus runs vehicles up to ~21h
  - allow_relief_points=True: forces VSP to cap by BLOCK span, not driver shift
  - unknown_deadhead_minutes: realistic proxy for cross-terminal repositioning

Usage: venv/bin/python scratch/compare_optibus.py <xlsx> --out <json> [--budget S]
                                                  [--config fair|naive]
"""
import sys, os, time, math, json, argparse, re, csv
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
os.environ.setdefault("INTERNAL_OPTIMIZER_KEY", "test-strong-key-for-pytest-32chars-ok")
os.environ.setdefault("DATABASE_URL", "sqlite:///./test.db")
os.environ.setdefault("REDIS_URL", "redis://localhost:6379/0")

import openpyxl
from src.services.optimizer_service import OptimizerService
from src.domain.models import Trip, VehicleType, AlgorithmType

ALGOS = [
    "greedy", "genetic", "simulated_annealing", "tabu_search",
    "mcnf", "assignment_vsp", "alns", "branch_and_price",
    "set_partitioning", "hybrid_pipeline", "joint_solver",
    "vcsp_pulp", "joint_bp", "regional", "lagrangean_joint",
    "bundle_method", "joint_timetable",
]


def _tmin(s):
    s = str(s).strip()
    if ":" in s:
        h, m = s.split(":")[:2]
        return int(h) * 60 + int(m)
    try:
        v = float(s)
        return int(v * 1440) if v < 2.5 else int(v)
    except Exception:
        return 0


def _norm(t):
    return t + 1440 if t < 180 else t


def load_stops():
    stops = {}
    path = os.path.join(os.path.dirname(__file__), "../../backend/src/modules/gtfs/fixtures/sunt_salvador/stops.txt")
    if not os.path.exists(path):
        path = os.path.join(os.path.dirname(__file__), "../backend/src/modules/gtfs/fixtures/sunt_salvador/stops.txt")
    if not os.path.exists(path):
        return stops
    with open(path, mode="r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            sid = row["stop_id"].strip()
            stops[sid] = (float(row["stop_lat"]), float(row["stop_lon"]))
    return stops


def load_excel_deadhead_matrix(path):
    from collections import defaultdict
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Duties"]
    rows = list(ws.iter_rows(values_only=True))
    H = {h: i for i, h in enumerate(rows[0])}
    dh_rows = [r for r in rows[1:] if r[H["Event Type"]] == "deadhead"]
    
    matrix = defaultdict(list)
    for r in dh_rows:
        o = int(r[H["Origin Stop Id"]]) if r[H["Origin Stop Id"]] is not None else None
        d = int(r[H["Destination Stop Id"]]) if r[H["Destination Stop Id"]] is not None else None
        if o is None or d is None:
            continue
        st = r[H["Start Time"]]
        et = r[H["End Time"]]
        dur = _tmin(et) - _tmin(st)
        if dur < 0: dur += 1440
        matrix[(o, d)].append(dur)
    wb.close()
    return {k: min(v) for k, v in matrix.items()}


def load_optibus(path, stops_map, excel_dh_matrix):
    """Return (trips, optibus_baseline_dict). Segments (_1,_2) merged into real trips."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Duties"]
    rows = list(ws.iter_rows(values_only=True))
    H = {h: i for i, h in enumerate(rows[0])}
    data = rows[1:]

    # ---- Optibus baseline (its own solution) ----
    blocks = {r[H["Vehicle Block Id"]] for r in data if r[H["Vehicle Block Id"]] is not None}
    duties = {r[H["Duty id"]] for r in data if r[H["Duty id"]] is not None}
    st_rows = [r for r in data if r[H["Event Type"]] == "service_trip"]
    dh_rows = [r for r in data if r[H["Event Type"]] == "deadhead"]
    baseline = {
        "vehicles": len(blocks),
        "duties": len(duties),
        "service_trip_rows": len(st_rows),
        "deadhead_events": len(dh_rows),
        "lines": sorted({str(r[H["Sign"]]) for r in st_rows}),
    }

    # ---- build merged real trips ----
    recs = []
    for r in st_rows:
        s = _norm(_tmin(r[H["Start Time"]]))
        e = _norm(_tmin(r[H["End Time"]]))
        if e < s:
            e += 1440
        raw = str(r[H["Trip Id"]]).strip()
        base = re.sub(r"_\d+$", "", raw)
        seg = re.search(r"_(\d+)$", raw)
        seg = int(seg.group(1)) if seg else 1
        recs.append({
            "base": base, "seg": seg, "s": s, "e": e,
            "o": r[H["Origin Stop Id"]], "d": r[H["Destination Stop Id"]],
            "dist": r[H["Distance"]], "sign": r[H["Sign"]],
            "dir": r[H["Direction"]],
        })
    wb.close()

    groups = {}
    for rec in recs:
        groups.setdefault(rec["base"], []).append(rec)

    trips = []
    tid = 1
    for base, segs in groups.items():
        segs.sort(key=lambda x: x["seg"])
        s0, eN = segs[0]["s"], segs[-1]["e"]
        o = segs[0]["o"]; d = segs[-1]["d"]
        dist = sum(float(x["dist"]) for x in segs if x["dist"] is not None)
        sign = segs[0]["sign"]
        try:
            line_id = int(sign)
        except Exception:
            line_id = 1
        o_id = int(o) if o is not None else 1
        d_id = int(d) if d is not None else 2

        # Look up coordinates
        o_lat, o_lon = stops_map.get(str(o_id), (None, None))
        d_lat, d_lon = stops_map.get(str(d_id), (None, None))

        # Pre-populate deadhead times using excel matrix
        dh_times = {}
        for (src, dest), dur in excel_dh_matrix.items():
            if src == d_id:
                dh_times[int(dest)] = int(dur)

        trips.append(Trip(
            id=tid, line_id=line_id,
            start_time=int(s0), end_time=int(eN), duration=int(eN - s0),
            origin_id=o_id, destination_id=d_id,
            distance_km=float(dist) if dist else 1.0,
            trip_group_id=tid, segment_count=len(segs),
            origin_latitude=o_lat, origin_longitude=o_lon,
            destination_latitude=d_lat, destination_longitude=d_lon,
            deadhead_times=dh_times
        ))
        tid += 1
    return trips, baseline


def max_concurrency(trips):
    ev = []
    for t in trips:
        ev.append((t.start_time, 1)); ev.append((t.end_time, -1))
    ev.sort(key=lambda e: (e[0], e[1]))
    cur = peak = 0
    for _, d in ev:
        cur += d; peak = max(peak, cur)
    return peak


def block_overlaps(block):
    ts = sorted(block.trips, key=lambda t: t.start_time)
    return sum(1 for a, b in zip(ts, ts[1:]) if b.start_time < a.end_time)


def make_params(config):
    if config == "fair":
        return {
            "min_layover_minutes": 0,
            "max_block_span_minutes": 1440,
            "max_vehicle_shift_minutes": 1440,
            "allow_relief_points": True,
            # deadhead inferido da própria timetable (tempos terminal->terminal);
            # pares sem rota de serviço continuam proibidos (999999) por padrão.
            "infer_deadhead_from_timetable": True,
            "enforce_min_interval": False,
            "strict_zero_gap_validation": False,
            "allow_multi_line_block": True,
            "disable_scale_decomposition": False,
            "fallback_deadhead_speed_kmh": 36.0,
            "fallback_deadhead_floor_minutes": 5,
        }
    # naive = old defaults that produced the 42-vehicle Mussurunga result
    return {
        "min_layover_minutes": 8,
        "max_vehicle_shift_minutes": 560,
        "unknown_deadhead_minutes": 15,
        "enforce_min_interval": False,
        "disable_scale_decomposition": False,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--out", required=True)
    ap.add_argument("--budget", type=float, default=30.0)
    ap.add_argument("--config", default="fair", choices=["fair", "naive"])
    args = ap.parse_args()

    stops_map = load_stops()
    excel_dh_matrix = load_excel_deadhead_matrix(args.xlsx)
    trips, baseline = load_optibus(args.xlsx, stops_map, excel_dh_matrix)
    lb = max_concurrency(trips)
    vt = [VehicleType(id=1, name="Convencional", passenger_capacity=80,
                      fixed_cost=1000.0, cost_per_km=3.0)]
    vsp_params = make_params(args.config)
    svc = OptimizerService()

    print(f"INSTANCE {os.path.basename(args.xlsx)[:30]} | merged_trips={len(trips)} "
          f"concurrency_LB={lb} | OPTIBUS veh={baseline['vehicles']} duties={baseline['duties']}")
    print(f"config={args.config} budget={args.budget}s\n")
    print(f"{'algo':<20}{'s':>7}{'veh':>6}{'cov':>11}{'ovl':>5}{'duties':>7}{'cost':>14}  status")
    print("-" * 80)

    results = []
    for algo in ALGOS:
        t0 = time.perf_counter()
        try:
            res = svc.run(trips=trips, vehicle_types=vt, algorithm=AlgorithmType(algo),
                          time_budget_s=args.budget, vsp_params=dict(vsp_params), cct_params={})
            el = time.perf_counter() - t0
            blocks = res.vsp.blocks if res.vsp else []
            covered = sum(len(b.trips) for b in blocks)
            uncov = len(res.vsp.unassigned_trips) if (res.vsp and res.vsp.unassigned_trips is not None) else -1
            overlaps = sum(block_overlaps(b) for b in blocks)
            cost = res.total_cost or 0.0
            ndut = len(res.csp.duties) if res.csp else 0
            ok = (uncov == 0 and overlaps == 0 and covered == len(trips)
                  and cost > 0 and math.isfinite(cost))
            status = "OK" if ok else f"FAIL cov={covered}/{len(trips)} unc={uncov} ovl={overlaps}"
            print(f"{algo:<20}{el:>7.1f}{len(blocks):>6}{covered:>8}/{len(trips):<3}"
                  f"{overlaps:>5}{ndut:>7}{cost:>14,.0f}  {status}")
            results.append({"algo": algo, "elapsed_s": round(el, 2), "vehicles": len(blocks),
                            "covered": covered, "unassigned": uncov, "overlaps": overlaps,
                            "duties": ndut, "cost": round(cost, 2), "status": status})
        except Exception as e:
            el = time.perf_counter() - t0
            print(f"{algo:<20}{el:>7.1f}  ERROR: {type(e).__name__}: {str(e)[:60]}")
            results.append({"algo": algo, "elapsed_s": round(el, 2), "vehicles": None,
                            "status": f"ERROR: {type(e).__name__}: {str(e)[:120]}"})

    feasible = [r for r in results if r.get("status") == "OK"]
    best = min(feasible, key=lambda r: (r["vehicles"], r["cost"])) if feasible else None
    out = {
        "instance": os.path.basename(args.xlsx),
        "config": args.config,
        "merged_trips": len(trips),
        "concurrency_lb": lb,
        "optibus": baseline,
        "results": results,
        "best_feasible": best,
    }
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)
    print(f"\nBEST OTIMIZ: {best['algo'] if best else None} "
          f"veh={best['vehicles'] if best else '-'} vs OPTIBUS {baseline['vehicles']}")
    print(f"WROTE {args.out}")


if __name__ == "__main__":
    main()
